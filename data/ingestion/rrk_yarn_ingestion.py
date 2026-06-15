"""
Ingest yarn purchase and goods receipt data from RRK Cottons India Excel export.

Run:
  python data/ingestion/rrk_yarn_ingestion.py --file /path/to/Yarn_Against_Order.xlsx
"""

from __future__ import annotations

import argparse
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from data.ingestion._env import load_project_env
from database.base import SessionLocal, is_duplicate_row, mark_latest
from database.ingestion_context import IngestionContext
from database.models.yarn_fabric import Yarn
from database.yarn_usd_prices import update_yarn_usd_prices
from database.validation.ingestion_validators import validate_yarn_price_inr

load_project_env()

logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)

SOURCE_NAME = "rrk_yarn_excel"
SCRIPT_VERSION = "rrk-yarn-v1.0"
SHEET_NAME = "PurGrnStatementReport"
EXCEL_EPOCH = date(1899, 12, 30)
MAX_REC_QTY_KG = Decimal("150000")

HEADER_ALIASES: dict[str, list[str]] = {
    "s_no": ["S.NO", "S NO", "SNO", "SR NO", "SR.NO"],
    "po_no": ["PO NO", "PO NO.", "PONO", "PO NUMBER"],
    "ref_no": ["REF NO", "REF NO.", "REFNO", "REF NUMBER"],
    "grn": [
        "GRN NO/DT/SUPPLIER/ORD TYPE",
        "GRN NO/DT/SUPPLIER/ORD TYPE ",
        "GRN NO / DT / SUPPLIER / ORD TYPE",
    ],
    "particulars": ["PARTICULARS", "PARTICULAR", "DESCRIPTION"],
    "rec_qty": ["REC QTY", "REC.QTY", "RECEIVED QTY", "REC QTY."],
    "po_rate": ["PO RATE", "PO.RATE", "P.O. RATE"],
    "rate": ["RATE"],
    "amount": ["AMOUNT", "AMT", "VALUE"],
    "dc_no": ["DC NO", "DC NO.", "DCNO", "DC NUMBER"],
    "dc_date": ["DC DATE", "DC.DATE", "D.C. DATE"],
}

FIBRE_PATTERN = re.compile(
    r"(\d+(?:\.\d+)?)\s*%\s*(COTTON|POLY(?:STER|ESTER)?|MODAL|VISCOSE|SPANDEX)\b",
    re.IGNORECASE,
)
YARN_COUNT_PATTERN = re.compile(r"(\d+)\s*s\b", re.IGNORECASE)
NON_DATE_DC_PATTERN = re.compile(
    r"[a-zA-Z]|&|\d+-[a-zA-Z]|[a-zA-Z]+\d+|\d+\s+[a-zA-Z]",
    re.IGNORECASE,
)
MONTH_PATTERN = re.compile(
    r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b",
    re.IGNORECASE,
)

DC_DATE_FORMATS = (
    "%d-%b-%y",
    "%d-%b-%Y",
    "%d/%m/%Y",
    "%d/%m/%y",
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d-%m-%y",
)


@dataclass
class ParsedParticulars:
    yarn_count: Optional[int] = None
    cotton: Decimal = Decimal("0")
    polyester: Decimal = Decimal("0")
    modal: Decimal = Decimal("0")
    viscose: Decimal = Decimal("0")
    spandex: Decimal = Decimal("0")
    fibre_type: str = "unknown"
    spinning_method: Optional[str] = None
    colour: str = "other"
    is_melange: bool = False
    is_recycled: bool = False
    is_bci: bool = False
    requires_review: bool = False
    data_notes: Optional[str] = None
    fibre_pct_suspicious: bool = False


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().upper())


def _cell_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _has_cell_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _effective_rate_from_row(row: dict[str, Any]) -> Optional[Decimal]:
    """Actual RATE paid takes precedence over agreed PO RATE."""
    actual_rate = _parse_decimal(row.get("rate"))
    agreed_po_rate = _parse_decimal(row.get("po_rate"))
    return actual_rate if actual_rate is not None else agreed_po_rate


def _is_empty_row(values: list[Any]) -> bool:
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in values
    )


def _is_header_row(values: list[Any]) -> bool:
    joined = " ".join(_normalize_header(v) for v in values if v is not None)
    return "S.NO" in joined or "PO NO" in joined


def _is_footer_row(row: dict[str, Any]) -> bool:
    particulars = _cell_text(row.get("particulars")) or ""
    po_no = _cell_text(row.get("po_no")) or ""
    if "TOTAL" in particulars.upper():
        return True
    if "TOTAL" in po_no.upper():
        return True
    return False


def _build_column_map(header_values: list[Any]) -> dict[str, int]:
    normalized_headers = [_normalize_header(value) for value in header_values]
    column_map: dict[str, int] = {}
    for field_name, aliases in HEADER_ALIASES.items():
        for idx, header in enumerate(normalized_headers):
            if header in {_normalize_header(alias) for alias in aliases}:
                column_map[field_name] = idx
                break
    return column_map


def read_excel_rows(file_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(file_path, read_only=False, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {file_path}")
    ws = wb[SHEET_NAME]

    raw_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if _is_empty_row(values):
            continue
        raw_rows.append(values)

    header_idx = next(
        (idx for idx, values in enumerate(raw_rows) if _is_header_row(values)),
        None,
    )
    if header_idx is None:
        raise ValueError("Header row containing 'S.NO' or 'PO NO' not found")

    column_map = _build_column_map(raw_rows[header_idx])
    required = {"po_no", "particulars", "rec_qty", "grn"}
    missing = sorted(required - set(column_map))
    if missing:
        raise ValueError(f"Required columns missing from Excel header: {missing}")

    rows: list[dict[str, Any]] = []
    for values in raw_rows[header_idx + 1 :]:
        if _is_empty_row(values):
            continue
        row = {
            field_name: values[idx] if idx < len(values) else None
            for field_name, idx in column_map.items()
        }
        if _is_footer_row(row):
            continue
        rows.append(row)
    return rows


def forward_fill_identifiers(rows: list[dict[str, Any]]) -> tuple[int, int, int, int]:
    po_fill_count = 0
    ref_fill_count = 0
    po_rate_fill_count = 0
    rate_fill_count = 0
    current_po: Optional[str] = None
    current_ref: Optional[str] = None
    current_po_rate: Any = None
    current_rate: Any = None

    for row in rows:
        po_raw = _cell_text(row.get("po_no"))
        if po_raw and po_raw.upper().startswith("RRK-"):
            current_po = po_raw
            current_ref = None
            current_po_rate = None
            current_rate = None
            row["po_no"] = current_po
        elif current_po and (po_raw is None or not po_raw):
            row["po_no"] = current_po
            po_fill_count += 1
        elif po_raw:
            row["po_no"] = po_raw

        ref_raw = _cell_text(row.get("ref_no"))
        if ref_raw:
            current_ref = ref_raw
            row["ref_no"] = current_ref
        elif current_ref and (ref_raw is None or not ref_raw):
            row["ref_no"] = current_ref
            ref_fill_count += 1

        po_rate_raw = row.get("po_rate")
        if _has_cell_value(po_rate_raw):
            current_po_rate = po_rate_raw
        elif current_po_rate is not None:
            row["po_rate"] = current_po_rate
            po_rate_fill_count += 1

        rate_raw = row.get("rate")
        if _has_cell_value(rate_raw):
            current_rate = rate_raw
        elif current_rate is not None:
            row["rate"] = current_rate
            rate_fill_count += 1

    return po_fill_count, ref_fill_count, po_rate_fill_count, rate_fill_count


def parse_grn_field(grn_raw: Any) -> tuple[Optional[str], Optional[date], Optional[str]]:
    text = _cell_text(grn_raw)
    if not text:
        return None, None, None

    parts = [part.strip() for part in text.split(" - ")]
    grn_number = parts[0] if parts else None
    grn_date: Optional[date] = None
    supplier: Optional[str] = None

    if len(parts) > 1:
        grn_date_str = parts[1]
        try:
            grn_date = datetime.strptime(grn_date_str, "%d-%b-%y").date()
        except ValueError:
            logger.warning("GRN date unparseable: %r — stored as NULL", grn_date_str)

    if len(parts) > 2:
        supplier = re.sub(r"\s+", " ", parts[2]).strip()
        supplier = re.sub(r"\s*-\s*[A-Z]$", "", supplier).strip()

    return grn_number, grn_date, supplier


def _map_fibre_keyword(keyword: str) -> Optional[str]:
    upper = keyword.upper()
    if upper == "COTTON":
        return "cotton"
    if upper in {"POLY", "POLYSTER", "POLYESTER"}:
        return "polyester"
    if upper == "MODAL":
        return "modal"
    if upper == "VISCOSE":
        return "viscose"
    if upper == "SPANDEX":
        return "spandex"
    return None


def _derive_fibre_type(
    cotton: Decimal,
    polyester: Decimal,
    modal: Decimal,
    viscose: Decimal,
    spandex: Decimal,
) -> str:
    if cotton == 100 and polyester == 0 and modal == 0 and viscose == 0 and spandex == 0:
        return "cotton"
    if cotton > 0 and polyester > 0 and viscose == 0 and modal == 0 and spandex == 0:
        return "cotton_polyester_blend"
    if cotton > 0 and modal > 0:
        return "cotton_modal_blend"
    if polyester > 0 and viscose > 0:
        return "polyester_viscose_blend"
    if modal == 100:
        return "modal"
    if polyester == 100:
        return "polyester"
    if cotton > 0 and spandex > 0:
        return "cotton_spandex_blend"
    if polyester > 0 and spandex > 0:
        return "polyester_spandex_blend"
    return "unknown"


def _parse_spinning_method(text_upper: str) -> Optional[str]:
    if "SEMI COMBED" in text_upper:
        return "semi_combed"
    if "COMBED" in text_upper:
        return "combed"
    if "OPEN END" in text_upper:
        return "open_end"
    if "RECYCLED" in text_upper:
        return "recycled"
    if "VORTEX" in text_upper:
        return "vortex"
    if "SLUB" in text_upper:
        return "slub"
    return None


def _parse_colour(text_upper: str) -> str:
    colour_rules = (
        ("LIGHT DENIM MELANGE", "light_denim_melange"),
        ("DARK DENIM MELANGE", "dark_denim_melange"),
        ("DENIM MELANGE", "denim_melange"),
        ("DARK GREY MELANGE", "dark_grey_melange"),
        ("LIGHT GREY MELANGE", "light_grey_melange"),
        ("MID GREY MELANGE", "mid_grey_melange"),
        ("MELANGE", "melange"),
        ("GREY", "grey"),
        ("BLACK", "black"),
        ("WHITE", "white"),
    )
    for token, colour in colour_rules:
        if token in text_upper:
            return colour
    return "other"


def parse_particulars(particulars: str, rate: Optional[Decimal] = None) -> ParsedParticulars:
    text_upper = particulars.upper()
    result = ParsedParticulars()

    count_matches = list(YARN_COUNT_PATTERN.finditer(particulars))
    if count_matches:
        result.yarn_count = int(count_matches[-1].group(1))

    normalized = re.sub(r"BCI\s+COTTON", "COTTON", text_upper)
    fibres = {
        "cotton": Decimal("0"),
        "polyester": Decimal("0"),
        "modal": Decimal("0"),
        "viscose": Decimal("0"),
        "spandex": Decimal("0"),
    }
    for match in FIBRE_PATTERN.finditer(normalized):
        pct = Decimal(str(match.group(1)))
        mapped = _map_fibre_keyword(match.group(2))
        if mapped:
            fibres[mapped] += pct

    result.cotton = fibres["cotton"]
    result.polyester = fibres["polyester"]
    result.modal = fibres["modal"]
    result.viscose = fibres["viscose"]
    result.spandex = fibres["spandex"]

    total_pct = sum(fibres.values())
    if total_pct > Decimal("110") or (Decimal("0") < total_pct < Decimal("90")):
        result.fibre_pct_suspicious = True
        result.requires_review = True

    result.fibre_type = _derive_fibre_type(
        result.cotton,
        result.polyester,
        result.modal,
        result.viscose,
        result.spandex,
    )
    result.spinning_method = _parse_spinning_method(text_upper)
    result.colour = _parse_colour(text_upper)
    result.is_melange = "MELANGE" in text_upper
    result.is_recycled = "RECYCLED" in text_upper
    result.is_bci = "BCI" in text_upper

    if text_upper.startswith("INTERLOCK FLEECE"):
        result.fibre_type = "fabric_not_yarn"
        result.requires_review = True
        result.data_notes = (
            "INTERLOCK FLEECE — this appears to be fabric, not yarn. "
            "Review before using in cost calculations."
        )

    if (
        "SPANDEX" in text_upper
        and result.yarn_count is None
        and "DIA" in text_upper
        and result.fibre_type != "fabric_not_yarn"
    ):
        result.fibre_type = "fabric_not_yarn"
        result.requires_review = True
        rate_text = str(rate) if rate is not None else "unknown"
        result.data_notes = (
            "Appears to be fabric/knit priced per metre, not yarn per kg. "
            f"Rate of INR {rate_text} may not be per kg. Review before use."
        )

    return result


def _looks_like_dc_number_not_date(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return True
    if stripped.lower() in {"nokia"}:
        return True
    if NON_DATE_DC_PATTERN.search(stripped) and not MONTH_PATTERN.search(stripped):
        return True
    return False


def parse_dc_date(dc_date_raw: Any) -> Optional[date]:
    from datetime import datetime as dt_class

    if isinstance(dc_date_raw, dt_class):
        return dc_date_raw.date()
    if dc_date_raw is None:
        return None
    if isinstance(dc_date_raw, str) and not dc_date_raw.strip():
        return None

    if isinstance(dc_date_raw, (int, float)) and not isinstance(dc_date_raw, bool):
        serial = int(dc_date_raw)
        if serial <= 0:
            return None
        return EXCEL_EPOCH + timedelta(days=serial)

    text = str(dc_date_raw).strip()
    if _looks_like_dc_number_not_date(text):
        logger.warning("DC DATE unparseable: %r — stored as NULL", dc_date_raw)
        return None

    for fmt in DC_DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    logger.warning("DC DATE unparseable: %r — stored as NULL", dc_date_raw)
    return None


def _parse_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        if isinstance(value, float) and value != value:
            return None
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _yarn_type_key(parsed: ParsedParticulars) -> Optional[str]:
    if parsed.fibre_type == "fabric_not_yarn":
        return None
    if parsed.is_recycled or parsed.spinning_method == "recycled":
        return "recycled_poly_cotton"
    if parsed.fibre_type == "modal":
        return "modal"
    if parsed.fibre_type == "cotton_polyester_blend":
        return "cotton_poly_blend"
    if parsed.fibre_type == "cotton" and parsed.yarn_count is not None:
        if parsed.yarn_count <= 27:
            return "cotton_25s"
        if parsed.yarn_count <= 32:
            return "cotton_30s"
        if parsed.yarn_count <= 36:
            return "cotton_34s"
        return "cotton_40s"
    return None


def _validate_rate(
    rate: Decimal,
    yarn_type_key: Optional[str],
    ctx: IngestionContext,
) -> tuple[Optional[Decimal], bool]:
    if yarn_type_key is None:
        return rate, False

    is_valid, reason = validate_yarn_price_inr(float(rate), yarn_type_key)
    if not is_valid:
        ctx.record_flag(f"rate validation: {reason}")
        return None, True
    return rate, False


def resolve_price_per_kg(
    po_rate: Optional[Decimal],
    parsed: ParsedParticulars,
) -> tuple[Optional[Decimal], bool]:
    """Derive price_per_kg from PO rate and parsed particulars (no ingestion log)."""
    if parsed.fibre_type == "fabric_not_yarn":
        return po_rate, False
    if po_rate is None:
        return None, False

    yarn_type_key = _yarn_type_key(parsed)
    if yarn_type_key is None:
        return po_rate, False

    is_valid, _ = validate_yarn_price_inr(float(po_rate), yarn_type_key)
    if not is_valid:
        return None, True
    return po_rate, False


def build_yarn_data_notes(
    *,
    source_file_name: str,
    yarn_type_raw: str,
    supplier_name: Optional[str],
    grn_number: Optional[str],
    grn_date: Optional[date],
    po_number: Optional[str],
    buyer_reference: Optional[str],
    quantity_kg: Optional[Decimal],
    price_per_kg: Optional[Decimal],
    fibre_type: str,
    requires_review: bool,
) -> str:
    notes = [
        f"Source: RRK Cottons, Tirupur — {source_file_name}",
        f"Yarn: {yarn_type_raw}",
        f"Supplier: {supplier_name}",
        f"GRN: {grn_number} dated {grn_date}",
        f"PO: {po_number}",
        f"Buyer ref: {buyer_reference}",
        f"Quantity: {quantity_kg} kg at INR {price_per_kg}/kg",
        "Currency: INR — price_per_kg is in Indian Rupees",
    ]
    if fibre_type == "fabric_not_yarn":
        notes.append("FLAG: This row appears to be fabric not yarn — review before use")
    if requires_review:
        notes.append("FLAG: Requires review — see above")
    return " | ".join(notes)


def ingest_rrk_yarn(file_path: Path, db: Session) -> dict[str, Any]:
    rows = read_excel_rows(file_path)
    po_filled, ref_filled, po_rate_filled, rate_filled = forward_fill_identifiers(rows)
    logger.info(
        "Forward-filled %s PO NO, %s REF NO, %s PO RATE, %s RATE values",
        po_filled,
        ref_filled,
        po_rate_filled,
        rate_filled,
    )

    requires_review_count = 0
    fabric_row_count = 0
    type_counts: dict[str, int] = {}
    supplier_counts: dict[str, int] = {}
    buyer_counts: dict[str, int] = {}
    date_values: list[date] = []

    with IngestionContext(
        source_name=SOURCE_NAME,
        script_version=SCRIPT_VERSION,
        data_source_url=str(file_path),
        db=db,
    ) as ctx:
        for row in rows:
            po_no = _cell_text(row.get("po_no"))
            ref_no = _cell_text(row.get("ref_no"))
            particulars = _cell_text(row.get("particulars"))
            rec_qty = _parse_decimal(row.get("rec_qty"))
            agreed_po_rate = _parse_decimal(row.get("po_rate"))
            effective_rate = _effective_rate_from_row(row)
            amount = _parse_decimal(row.get("amount"))
            dc_no = _cell_text(row.get("dc_no"))
            dc_date = parse_dc_date(row.get("dc_date"))

            if rec_qty is None or rec_qty == 0:
                ctx.rejected(f"REC QTY missing or zero: {row.get('rec_qty')!r}")
                continue
            if rec_qty > MAX_REC_QTY_KG:
                ctx.rejected(f"REC QTY implausibly large: {rec_qty}")
                continue
            if not particulars:
                ctx.rejected("PARTICULARS missing or blank")
                continue
            if not po_no or not po_no.upper().startswith("RRK-"):
                ctx.rejected(f"PO NO invalid after forward fill: {po_no!r}")
                continue

            grn_number, grn_date, supplier = parse_grn_field(row.get("grn"))
            parsed = parse_particulars(particulars, rate=effective_rate)

            rate_validation_failed = False
            price_per_kg: Optional[Decimal] = None
            if parsed.fibre_type == "fabric_not_yarn":
                price_per_kg = effective_rate
            elif effective_rate is not None:
                yarn_type_key = _yarn_type_key(parsed)
                price_per_kg, rate_validation_failed = _validate_rate(
                    effective_rate, yarn_type_key, ctx
                )
                if price_per_kg is None and effective_rate is not None and yarn_type_key is None:
                    price_per_kg = effective_rate

            requires_review = (
                parsed.requires_review
                or parsed.fibre_pct_suspicious
                or rate_validation_failed
            )
            if requires_review:
                requires_review_count += 1
            if parsed.fibre_type == "fabric_not_yarn":
                fabric_row_count += 1

            as_of_date = grn_date or date.today()
            date_values.append(as_of_date)

            if is_duplicate_row(
                db,
                Yarn,
                filter_kwargs={
                    "grn_number": grn_number,
                    "buyer_reference": ref_no,
                },
                value_kwargs={
                    "quantity_kg": rec_qty,
                    "price_per_kg": price_per_kg,
                },
            ):
                ctx.stale()
                continue

            mark_latest(
                db,
                Yarn,
                {"grn_number": grn_number, "po_number": po_no},
            )

            data_notes = build_yarn_data_notes(
                source_file_name=file_path.name,
                yarn_type_raw=particulars,
                supplier_name=supplier,
                grn_number=grn_number,
                grn_date=grn_date,
                po_number=po_no,
                buyer_reference=ref_no,
                quantity_kg=rec_qty,
                price_per_kg=price_per_kg,
                fibre_type=parsed.fibre_type,
                requires_review=requires_review,
            )

            yarn_row = Yarn(
                fibre_type=parsed.fibre_type,
                fibre_content_pct_cotton=parsed.cotton,
                fibre_content_pct_polyester=parsed.polyester,
                fibre_content_pct_modal=parsed.modal,
                fibre_content_pct_viscose=parsed.viscose,
                fibre_content_pct_spandex=parsed.spandex,
                count=f"{parsed.yarn_count}s" if parsed.yarn_count is not None else None,
                spinning_method=parsed.spinning_method,
                colour=parsed.colour,
                is_melange=parsed.is_melange,
                is_recycled=parsed.is_recycled,
                is_bci=parsed.is_bci,
                requires_review=requires_review,
                origin_city="Tirupur",
                origin_country="India",
                price_per_kg=price_per_kg,
                price_per_kg_usd=None,
                local_currency="INR",
                supplier_name=supplier,
                buyer_reference=ref_no,
                po_number=po_no,
                grn_number=grn_number,
                grn_date=grn_date,
                quantity_kg=rec_qty,
                po_rate_inr=agreed_po_rate,
                amount_inr=amount,
                dc_number=dc_no,
                dc_date=dc_date,
                yarn_type_raw=particulars,
                availability_signal="available",
                confidence_score=Decimal("0.90"),
                as_of_date=as_of_date,
                source="rrk_excel_import",
                data_source_url=str(file_path),
                data_notes=data_notes,
                pulled_at=datetime.now(timezone.utc),
                is_latest=True,
            )
            db.add(yarn_row)
            ctx.inserted()

            type_counts[parsed.fibre_type] = type_counts.get(parsed.fibre_type, 0) + 1
            supplier_key = supplier or "UNKNOWN"
            supplier_counts[supplier_key] = supplier_counts.get(supplier_key, 0) + 1
            buyer_key = ref_no or "UNKNOWN"
            buyer_counts[buyer_key] = buyer_counts.get(buyer_key, 0) + 1

        if date_values:
            ctx.set_as_of_date(min(date_values))

        db.commit()

        usd_updated = update_yarn_usd_prices(db, only_null=True)
        logger.info("Post-ingestion USD price update: %s row(s)", usd_updated)

        return {
            "ctx": ctx,
            "requires_review_count": requires_review_count,
            "fabric_row_count": fabric_row_count,
            "type_counts": type_counts,
            "supplier_counts": supplier_counts,
            "buyer_counts": buyer_counts,
            "min_date": min(date_values) if date_values else None,
            "max_date": max(date_values) if date_values else None,
            "usd_updated": usd_updated,
        }


def print_summary(summary: dict[str, Any]) -> None:
    ctx: IngestionContext = summary["ctx"]
    print("=== RRK YARN INGESTION COMPLETE ===")
    print(f"  Rows inserted:         {ctx.log.rows_inserted}")
    print(f"  Rows rejected:         {ctx.log.rows_rejected}")
    print(f"  Rows stale:            {ctx.log.rows_stale}")
    print(f"  Rows requiring review: {summary['requires_review_count']}")
    print(f"  Fabric rows flagged:   {summary['fabric_row_count']}")
    min_date = summary["min_date"]
    max_date = summary["max_date"]
    if min_date and max_date:
        print(f"  Date range:            {min_date} to {max_date}")
    print(f"  USD prices updated:    {summary.get('usd_updated', 0)}")
    print()
    print("  Yarn types:")
    for yarn_type, count in sorted(summary["type_counts"].items()):
        print(f"    {yarn_type:<35} {count} rows")
    print()
    print("  Suppliers:")
    for supplier, count in sorted(summary["supplier_counts"].items()):
        print(f"    {supplier:<45} {count} deliveries")
    print()
    print("  Buyers:")
    for buyer, count in sorted(summary["buyer_counts"].items()):
        print(f"    {buyer:<45} {count} rows")


def main() -> int:
    parser = argparse.ArgumentParser(description="Ingest RRK yarn Excel purchase/GRN data")
    parser.add_argument("--file", required=True, type=Path, help="Path to Yarn_Against_Order.xlsx")
    args = parser.parse_args()

    if not args.file.exists():
        raise FileNotFoundError(f"Excel file not found: {args.file}")

    db = SessionLocal()
    try:
        summary = ingest_rrk_yarn(args.file, db)
        print_summary(summary)
    finally:
        db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
