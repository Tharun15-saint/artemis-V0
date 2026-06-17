"""
World Bank Commodity Markets Outlook Pink Sheet (CMO Historical Data Monthly).

Reads the local Pink Sheet workbook and writes:
  - COTLOOK_A monthly observations to cotton_price_observation
  - Brent/WTI monthly spot prices to crude_oil

Pink Sheet cotton column is USD/kg; Artemis stores normalized cents/lb and USD/kg.
"""

from __future__ import annotations

import argparse
import logging
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from data.ingestion._env import load_project_env
from database.base import SessionLocal, is_duplicate_row, mark_latest
from database.ingestion_context import IngestionContext
from database.models import CrudeOil

load_project_env()

logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)

SCRIPT_VERSION = "1.0.0"
SOURCE_NAME = "world_bank_pink_sheet"
SHEET_NAME = "Monthly Prices"
WB_PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/CMO-Historical-Data-Monthly.xlsx"
)

COTTON_COL = "Cotton, A Index"
BRENT_COL = "Crude oil, Brent"
WTI_COL = "Crude oil, WTI"
MISSING_VALUES = {None, "", "…", "..", "."}

# 1 cent/lb = 0.022046 USD/kg  →  USD/kg to cents/lb = usd_kg / 0.022046
USD_KG_TO_CENTS_LB = Decimal("1") / Decimal("0.022046")


@dataclass(frozen=True)
class PinkSheetRow:
    as_of_date: date
    cotton_usd_kg: Optional[Decimal]
    brent_usd_bbl: Optional[Decimal]
    wti_usd_bbl: Optional[Decimal]


def quantize(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


def parse_period_code(code: str) -> date:
    year_str, month_str = code.split("M", 1)
    return date(int(year_str), int(month_str), 1)


def _as_decimal(value: Any) -> Optional[Decimal]:
    if value in MISSING_VALUES:
        return None
    try:
        return quantize(Decimal(str(float(value))))
    except (TypeError, ValueError):
        return None


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows):
        if row and row[0] is None and row[1] and "Crude oil, average" in str(row[1]):
            columns = {
                str(name).strip(): col_idx
                for col_idx, name in enumerate(row)
                if name not in MISSING_VALUES
            }
            return idx, columns
    raise ValueError("Could not locate Pink Sheet commodity header row")


def load_pink_sheet_rows(source_file: Path) -> list[PinkSheetRow]:
    workbook = load_workbook(source_file, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Sheet {SHEET_NAME!r} not found in {source_file}")

    sheet = workbook[SHEET_NAME]
    raw_rows = list(sheet.iter_rows(values_only=True))
    header_idx, columns = _find_header_row(raw_rows)

    required = (COTTON_COL, BRENT_COL, WTI_COL)
    missing = [name for name in required if name not in columns]
    if missing:
        raise ValueError(f"Missing Pink Sheet columns: {missing}")

    parsed: list[PinkSheetRow] = []
    for row in raw_rows[header_idx + 2 :]:
        period_code = row[0]
        if not period_code or not isinstance(period_code, str) or "M" not in period_code:
            continue

        cotton = _as_decimal(row[columns[COTTON_COL]])
        brent = _as_decimal(row[columns[BRENT_COL]])
        wti = _as_decimal(row[columns[WTI_COL]])
        if cotton is None and brent is None and wti is None:
            continue

        parsed.append(
            PinkSheetRow(
                as_of_date=parse_period_code(period_code),
                cotton_usd_kg=cotton,
                brent_usd_bbl=brent,
                wti_usd_bbl=wti,
            )
        )

    logger.info("Loaded %s Pink Sheet monthly rows from %s", len(parsed), source_file)
    return parsed


def get_series_id(db: Session, series_code: str) -> int:
    row = db.execute(
        text("SELECT series_id FROM cotton_price_series WHERE series_code=:code"),
        {"code": series_code},
    ).fetchone()
    if not row:
        raise ValueError(f"Series {series_code} not found in cotton_price_series")
    return row[0]


def demote_cotton_observation(db: Session, series_id: int, obs_date: date) -> None:
    db.execute(
        text(
            """
        UPDATE cotton_price_observation
        SET is_latest=0, updated_at=CURRENT_TIMESTAMP
        WHERE series_id=:sid AND as_of_date=:dt AND is_latest=1
    """
        ),
        {"sid": series_id, "dt": obs_date.isoformat()},
    )


def insert_cotlook_observation(
    db: Session,
    series_id: int,
    obs_date: date,
    cotton_usd_kg: Decimal,
    pulled_at: datetime,
) -> None:
    cents_per_lb = quantize(cotton_usd_kg * USD_KG_TO_CENTS_LB)
    if cents_per_lb < Decimal("10") or cents_per_lb > Decimal("400"):
        raise ValueError(f"COTLOOK_A price out of range on {obs_date}: {cents_per_lb}")

    demote_cotton_observation(db, series_id, obs_date)
    db.execute(
        text(
            """
        INSERT INTO cotton_price_observation
          (series_id, series_code, as_of_date, price_value, price_unit,
           price_in_usd_cents_per_lb, price_in_usd_per_kg,
           raw_value_original_unit, original_unit,
           source_document, source_url,
           data_quality, data_notes, is_estimate, is_latest, pulled_at,
           created_at, updated_at)
        VALUES
          (:sid, 'COTLOOK_A', :dt, :cents, 'cents_per_lb',
           :cents, :usd_kg,
           :raw_kg, 'USD/kg',
           'World Bank Pink Sheet — Cotton, A Index (Cotlook A)',
           :src_url,
           'verified',
           'World Bank CMO Historical Data Monthly nominal USD/kg converted to cents/lb',
           0, 1, :pulled,
           CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
    """
        ),
        {
            "sid": series_id,
            "dt": obs_date.isoformat(),
            "cents": float(cents_per_lb),
            "usd_kg": float(cotton_usd_kg),
            "raw_kg": float(cotton_usd_kg),
            "src_url": WB_PINK_SHEET_URL,
            "pulled": pulled_at.isoformat(),
        },
    )


def backfill_cotlook_a(
    db: Session,
    ctx: IngestionContext,
    rows: list[PinkSheetRow],
) -> dict[str, int]:
    series_id = get_series_id(db, "COTLOOK_A")
    pulled_at = datetime.now(timezone.utc)
    inserted = skipped = rejected = 0

    for row in rows:
        if row.cotton_usd_kg is None:
            rejected += 1
            ctx.increment_rejected(f"missing_cotton_{row.as_of_date.isoformat()}")
            continue

        existing = db.execute(
            text(
                """
            SELECT price_in_usd_per_kg
            FROM cotton_price_observation
            WHERE series_id=:sid AND as_of_date=:dt AND is_latest=1
        """
            ),
            {"sid": series_id, "dt": row.as_of_date.isoformat()},
        ).fetchone()

        if existing and Decimal(str(existing[0])) == row.cotton_usd_kg:
            skipped += 1
            ctx.stale()
            continue

        try:
            insert_cotlook_observation(db, series_id, row.as_of_date, row.cotton_usd_kg, pulled_at)
        except ValueError as exc:
            rejected += 1
            ctx.increment_rejected(str(exc))
            continue

        ctx.increment_inserted()
        inserted += 1

    db.commit()
    logger.info(
        "COTLOOK_A Pink Sheet backfill: inserted=%s skipped=%s rejected=%s",
        inserted,
        skipped,
        rejected,
    )
    return {"inserted": inserted, "skipped": skipped, "rejected": rejected}


def _days_since_refresh(as_of: date) -> int:
    return (date.today() - as_of).days


def append_crude_row(
    db: Session,
    ctx: IngestionContext,
    as_of: date,
    brent: Optional[Decimal],
    wti: Optional[Decimal],
    pulled_at: datetime,
) -> bool:
    if brent is None and wti is None:
        ctx.increment_rejected(f"crude_oil: no Brent or WTI on {as_of.isoformat()}")
        return False

    for label, price in (("brent", brent), ("wti", wti)):
        if price is not None and (price < Decimal("1") or price > Decimal("250")):
            ctx.increment_rejected(
                f"crude_oil: {label} {price} out of range on {as_of.isoformat()}"
            )
            return False

    value_kwargs = {
        "brent_spot": brent,
        "wti_spot": wti,
        "as_of_date": as_of,
        "source": SOURCE_NAME,
        "data_source_url": WB_PINK_SHEET_URL,
    }
    if is_duplicate_row(db, CrudeOil, {"as_of_date": as_of}, value_kwargs):
        ctx.stale()
        return True

    mark_latest(db, CrudeOil, {"as_of_date": as_of})
    db.add(
        CrudeOil(
            brent_spot=brent,
            wti_spot=wti,
            as_of_date=as_of,
            days_since_refresh=_days_since_refresh(as_of),
            aggregation_period="monthly",
            source=SOURCE_NAME,
            data_source_url=WB_PINK_SHEET_URL,
            refresh="monthly",
            pulled_at=pulled_at,
            is_latest=True,
        )
    )
    ctx.increment_inserted()
    return True


def backfill_crude_oil(
    db: Session,
    ctx: IngestionContext,
    rows: list[PinkSheetRow],
) -> dict[str, int]:
    pulled_at = datetime.now(timezone.utc)
    inserted = skipped = rejected = 0

    for row in rows:
        if row.brent_usd_bbl is None and row.wti_usd_bbl is None:
            rejected += 1
            ctx.increment_rejected(f"crude_oil: empty row {row.as_of_date.isoformat()}")
            continue

        before_inserted = ctx.log.rows_inserted if ctx.log else 0
        before_stale = ctx.log.rows_stale if ctx.log else 0
        before_rejected = ctx.log.rows_rejected if ctx.log else 0

        append_crude_row(
            db,
            ctx,
            row.as_of_date,
            row.brent_usd_bbl,
            row.wti_usd_bbl,
            pulled_at,
        )

        after_inserted = ctx.log.rows_inserted if ctx.log else 0
        after_stale = ctx.log.rows_stale if ctx.log else 0
        after_rejected = ctx.log.rows_rejected if ctx.log else 0

        if after_inserted > before_inserted:
            inserted += 1
        elif after_stale > before_stale:
            skipped += 1
        elif after_rejected > before_rejected:
            rejected += 1

    db.commit()
    logger.info(
        "Crude oil Pink Sheet backfill: inserted=%s skipped=%s rejected=%s",
        inserted,
        skipped,
        rejected,
    )
    return {"inserted": inserted, "skipped": skipped, "rejected": rejected}


def run_backfill(source_file: Path) -> bool:
    if not source_file.exists():
        logger.error("Pink Sheet file not found: %s", source_file)
        return False

    rows = load_pink_sheet_rows(source_file)
    db = SessionLocal()
    try:
        with IngestionContext(
            source_name=f"{SOURCE_NAME}_cotlook_a",
            script_version=SCRIPT_VERSION,
            data_source_url=WB_PINK_SHEET_URL,
            db=db,
        ) as ctx:
            ctx.set_as_of_date(date.today())
            cotlook_stats = backfill_cotlook_a(db, ctx, rows)

        with IngestionContext(
            source_name=f"{SOURCE_NAME}_crude_oil",
            script_version=SCRIPT_VERSION,
            data_source_url=WB_PINK_SHEET_URL,
            db=db,
        ) as ctx:
            ctx.set_as_of_date(date.today())
            crude_stats = backfill_crude_oil(db, ctx, rows)

        logger.info(
            """
Pink Sheet backfill complete:
  COTLOOK_A: inserted=%s skipped=%s rejected=%s
  crude_oil: inserted=%s skipped=%s rejected=%s
""",
            cotlook_stats["inserted"],
            cotlook_stats["skipped"],
            cotlook_stats["rejected"],
            crude_stats["inserted"],
            crude_stats["skipped"],
            crude_stats["rejected"],
        )
        return True
    except Exception as exc:
        logger.critical("Pink Sheet ingestion failed: %s", exc, exc_info=True)
        db.rollback()
        return False
    finally:
        db.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Ingest World Bank Pink Sheet cotton and crude oil history"
    )
    parser.add_argument(
        "--source-file",
        required=True,
        help="Path to CMO-Historical-Data-Monthly.xlsx",
    )
    args = parser.parse_args()
    return 0 if run_backfill(Path(args.source_file)) else 1


if __name__ == "__main__":
    raise SystemExit(main())
