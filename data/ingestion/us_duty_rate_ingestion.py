"""
USITC HTS duty rate ingestion — Chapter 61 apparel schedule + country effective rates.

Reads USITC HTS Revision 10 XLSX, loads us_duty_rate_schedule for Chapter 61
(indent 1–4 rows with a General Rate of Duty), then derives
us_duty_country_effective_rate for six Artemis origin countries.

Source: https://www.usitc.gov/harmonized_tariff_information
"""

from __future__ import annotations

import logging
import re
import shutil
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database.base import SessionLocal, mark_latest
from database.ingestion_context import IngestionContext
from database.models.trade import UsDutyCountryEffectiveRate, UsDutyRateSchedule

logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)

SOURCE_NAME = "usitc_hts_2026_rev10"
SCRIPT_VERSION = "1.0.0"
DATA_SOURCE_URL = "https://www.usitc.gov/harmonized_tariff_information"
HTS_REVISION = "2026_revision_10"
EFFECTIVE_DATE = date(2026, 5, 28)
LAST_VERIFIED = date(2026, 6, 15)
IEEPA_PCT = Decimal("10.0")
SECTION_301_CHINA_PCT = Decimal("7.5")
CHAPTER = 61

UPLOAD_PATH = Path("/mnt/user-data/uploads/hts_2026_revision_10_xls.xlsx")
PROJECT_REFERENCE_PATH = (
    Path(__file__).resolve().parents[1] / "reference" / "hts_2026_revision_10_xls.xlsx"
)

CAFTA_DR_CODES = frozenset({"DO", "GT", "HN", "NI", "SV"})

ORIGIN_COUNTRIES = (
    {"name": "Bangladesh", "iso2": "BD"},
    {"name": "India", "iso2": "IN"},
    {"name": "Vietnam", "iso2": "VN"},
    {"name": "Jordan", "iso2": "JO"},
    {"name": "China", "iso2": "CN"},
    {"name": "Sri Lanka", "iso2": "LK"},
)

JORDAN_NON_QUALIFYING_NOTE = (
    "Rate applies when yarn-forward rule is NOT met — third-country yarn "
    "(e.g. Indian yarn from RRK) used without QIZ compliance. Full NTR + 10% IEEPA applies."
)


def _jordan_qualifying_notes(ntr_rate_pct: Decimal) -> str:
    non_qual = ntr_rate_pct + IEEPA_PCT
    return (
        "JUSFTA 0% applies IF yarn-forward rule is met (yarn formed in Jordan or US). "
        f"If Indian or third-country yarn is used without QIZ/JUSFTA compliance, effective rate is "
        f"{non_qual}% (NTR + IEEPA). RRK sources yarn from India — verify qualifying industrial "
        "zone (QIZ) status or direct Jordan origin compliance before assuming 0%."
    )

_PERCENT_RE = re.compile(r"([\d.]+)\s*%")


def resolve_hts_workbook_path() -> Path:
    """Prefer uploaded file; copy into project reference directory if needed."""
    if UPLOAD_PATH.is_file():
        PROJECT_REFERENCE_PATH.parent.mkdir(parents=True, exist_ok=True)
        if not PROJECT_REFERENCE_PATH.is_file():
            shutil.copy2(UPLOAD_PATH, PROJECT_REFERENCE_PATH)
            logger.info("Copied HTS workbook to %s", PROJECT_REFERENCE_PATH)
        return UPLOAD_PATH
    if PROJECT_REFERENCE_PATH.is_file():
        return PROJECT_REFERENCE_PATH
    raise FileNotFoundError(
        f"HTS workbook not found at {UPLOAD_PATH} or {PROJECT_REFERENCE_PATH}"
    )


def _heading_from_hts(hts_number: str) -> str:
    digits = re.sub(r"\D", "", hts_number)
    return digits[:4]


def _special_country_codes(special_rate: Optional[str]) -> set[str]:
    if not special_rate:
        return set()
    match = re.search(r"\(([^)]+)\)", str(special_rate))
    if not match:
        return set()
    return {code.strip().upper() for code in match.group(1).split(",") if code.strip()}


def parse_general_rate(general_rate: Optional[str]) -> tuple[Optional[Decimal], str, bool]:
    """
    Parse General Rate of Duty.
    Returns (ntr_rate_pct, ntr_rate_text, ntr_rate_is_compound).
    """
    if general_rate is None:
        return None, "", False

    text = str(general_rate).strip()
    if not text:
        return None, "", False

    rate_text = text[:100]
    lower = text.lower()
    if lower == "free":
        return Decimal("0.0"), rate_text, False

    if "¢" in text or "/kg" in lower or "+" in text:
        return None, rate_text, True

    match = _PERCENT_RE.search(text)
    if match:
        return Decimal(match.group(1)), rate_text, False

    return None, rate_text, True


def _as_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_int(value) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    return int(value)


def load_chapter_61_schedule_rows(workbook_path: Path) -> list[dict]:
    """Read Chapter 61 HTS rows from the USITC workbook."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["HTS data export"] if "HTS data export" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows: list[dict] = []
    for raw in ws.iter_rows(values_only=True):
        hts_raw = raw[0]
        if hts_raw is None or str(hts_raw).strip() == "":
            continue

        hts_number = str(hts_raw).strip()
        if not hts_number.startswith("61"):
            continue

        indent_level = _as_int(raw[1])
        if indent_level is None or indent_level < 1 or indent_level > 4:
            continue

        general_rate = _as_text(raw[4])
        if general_rate is None:
            continue

        ntr_rate_pct, ntr_rate_text, ntr_is_compound = parse_general_rate(general_rate)
        special_rate = _as_text(raw[5])
        special_codes = _special_country_codes(special_rate)

        rows.append(
            {
                "hts_number": hts_number,
                "hts_description": _as_text(raw[2]),
                "chapter": CHAPTER,
                "heading": _heading_from_hts(hts_number),
                "indent_level": indent_level,
                "ntr_rate_pct": ntr_rate_pct,
                "ntr_rate_text": ntr_rate_text or general_rate[:100],
                "ntr_rate_is_compound": ntr_is_compound,
                "fta_free_countries": special_rate,
                "jusfta_jordan_free": "JO" in special_codes,
                "korus_korea_free": "KR" in special_codes,
                "morocco_fta_free": "MA" in special_codes,
                "cafta_dr_free": bool(special_codes & CAFTA_DR_CODES),
                "column2_rate_text": (_as_text(raw[6]) or "")[:100] or None,
                "additional_duties_text": (_as_text(raw[8]) or "")[:200] or None,
            }
        )

    wb.close()
    return rows


def ingest_schedule(db: Session, ctx: IngestionContext, schedule_rows: list[dict]) -> int:
    inserted = 0
    for row in schedule_rows:
        mark_latest(
            db,
            UsDutyRateSchedule,
            {"hts_number": row["hts_number"], "hts_revision": HTS_REVISION},
        )
        record = UsDutyRateSchedule(
            hts_number=row["hts_number"],
            hts_description=row["hts_description"],
            chapter=row["chapter"],
            heading=row["heading"],
            indent_level=row["indent_level"],
            ntr_rate_pct=row["ntr_rate_pct"],
            ntr_rate_text=row["ntr_rate_text"],
            ntr_rate_is_compound=row["ntr_rate_is_compound"],
            fta_free_countries=row["fta_free_countries"],
            jusfta_jordan_free=row["jusfta_jordan_free"],
            korus_korea_free=row["korus_korea_free"],
            morocco_fta_free=row["morocco_fta_free"],
            cafta_dr_free=row["cafta_dr_free"],
            column2_rate_text=row["column2_rate_text"],
            additional_duties_text=row["additional_duties_text"],
            section_301_china_applies=False,
            section_301_china_rate_pct=None,
            section_301_list=None,
            ieepa_universal_rate_pct=IEEPA_PCT,
            ieepa_universal_notes="IEEPA universal 10% ad valorem (default Artemis assumption)",
            effective_date=EFFECTIVE_DATE,
            hts_revision=HTS_REVISION,
            source=SOURCE_NAME,
            data_source_url=DATA_SOURCE_URL,
            last_verified=LAST_VERIFIED,
            is_latest=True,
        )
        db.add(record)
        ctx.inserted()
        inserted += 1

    db.commit()
    return inserted


def _effective_rate_filter(
    hts_number: str,
    origin_iso2: str,
    yarn_forward_met_assumption: str,
) -> dict:
    return {
        "hts_number": hts_number,
        "origin_iso2": origin_iso2,
        "yarn_forward_met_assumption": yarn_forward_met_assumption,
        "source": SOURCE_NAME,
    }


def _build_standard_effective_rate(
    schedule_row: dict,
    origin_iso2: str,
) -> tuple[Decimal, Optional[Decimal], Optional[Decimal], Optional[Decimal], Optional[str], str, bool, bool, str]:
    """Returns rate components for non-Jordan origins."""
    ntr = schedule_row["ntr_rate_pct"]
    fta_rate: Optional[Decimal] = None
    fta_program: Optional[str] = None
    section_301: Optional[Decimal] = None
    ieepa = IEEPA_PCT
    yarn_forward = False
    uflpa = False
    assumption = "assumed_met"

    if origin_iso2 == "CN":
        section_301 = SECTION_301_CHINA_PCT
        uflpa = True
        effective = ntr + SECTION_301_CHINA_PCT + IEEPA_PCT
        notes = "Section 301 List 2 apparel default 7.5% + IEEPA 10%"
        return effective, fta_rate, section_301, ieepa, fta_program, notes, yarn_forward, uflpa, assumption

    effective = ntr + IEEPA_PCT
    notes = "NTR + IEEPA 10%; no FTA or Section 301"
    return effective, fta_rate, section_301, ieepa, fta_program, notes, yarn_forward, uflpa, assumption


def _build_jordan_effective_rows(
    schedule_row: dict,
) -> list[dict]:
    """Jordan stores qualifying (assumed_met) and non-qualifying (not_met) scenarios."""
    ntr = schedule_row["ntr_rate_pct"]
    rows: list[dict] = []

    if schedule_row["jusfta_jordan_free"]:
        rows.append(
            {
                "origin_country": "Jordan",
                "origin_iso2": "JO",
                "ntr_rate_pct": ntr,
                "fta_rate_pct": Decimal("0.0"),
                "fta_program": "JUSFTA",
                "section_301_additional_pct": None,
                "ieepa_additional_pct": IEEPA_PCT,
                "effective_rate_pct": Decimal("0.0"),
                "effective_rate_notes": _jordan_qualifying_notes(ntr),
                "yarn_forward_required": True,
                "yarn_forward_met_assumption": "assumed_met",
                "uflpa_risk": False,
            }
        )
    else:
        rows.append(
            {
                "origin_country": "Jordan",
                "origin_iso2": "JO",
                "ntr_rate_pct": ntr,
                "fta_rate_pct": None,
                "fta_program": None,
                "section_301_additional_pct": None,
                "ieepa_additional_pct": IEEPA_PCT,
                "effective_rate_pct": ntr + IEEPA_PCT,
                "effective_rate_notes": _jordan_qualifying_notes(ntr),
                "yarn_forward_required": True,
                "yarn_forward_met_assumption": "assumed_met",
                "uflpa_risk": False,
            }
        )

    rows.append(
        {
            "origin_country": "Jordan (non-qualifying)",
            "origin_iso2": "JO",
            "ntr_rate_pct": ntr,
            "fta_rate_pct": None,
            "fta_program": None,
            "section_301_additional_pct": None,
            "ieepa_additional_pct": IEEPA_PCT,
            "effective_rate_pct": (ntr + IEEPA_PCT).quantize(Decimal("0.0001")),
            "effective_rate_notes": JORDAN_NON_QUALIFYING_NOTE,
            "yarn_forward_required": True,
            "yarn_forward_met_assumption": "not_met",
            "uflpa_risk": False,
        }
    )
    return rows


def ingest_effective_rates(db: Session, ctx: IngestionContext, schedule_rows: list[dict]) -> int:
    """Populate country effective rates for rows with numeric NTR rates."""
    numeric_rows = [row for row in schedule_rows if row["ntr_rate_pct"] is not None]
    inserted = 0

    for schedule_row in numeric_rows:
        for origin in ORIGIN_COUNTRIES:
            if origin["iso2"] == "JO":
                jordan_rows = _build_jordan_effective_rows(schedule_row)
                for row in jordan_rows:
                    mark_latest(
                        db,
                        UsDutyCountryEffectiveRate,
                        _effective_rate_filter(
                            schedule_row["hts_number"],
                            row["origin_iso2"],
                            row["yarn_forward_met_assumption"],
                        ),
                    )
                    db.add(
                        UsDutyCountryEffectiveRate(
                            hts_number=schedule_row["hts_number"],
                            origin_country=row["origin_country"],
                            origin_iso2=row["origin_iso2"],
                            ntr_rate_pct=row["ntr_rate_pct"],
                            fta_rate_pct=row["fta_rate_pct"],
                            fta_program=row["fta_program"],
                            section_301_additional_pct=row["section_301_additional_pct"],
                            ieepa_additional_pct=row["ieepa_additional_pct"],
                            effective_rate_pct=row["effective_rate_pct"],
                            effective_rate_notes=row["effective_rate_notes"],
                            yarn_forward_required=row["yarn_forward_required"],
                            yarn_forward_met_assumption=row["yarn_forward_met_assumption"],
                            uflpa_risk=row["uflpa_risk"],
                            as_of_date=EFFECTIVE_DATE,
                            source=SOURCE_NAME,
                            is_latest=True,
                        )
                    )
                    ctx.inserted()
                    inserted += 1
                continue

            (
                effective_rate,
                fta_rate,
                section_301,
                ieepa,
                fta_program,
                notes,
                yarn_forward,
                uflpa,
                assumption,
            ) = _build_standard_effective_rate(schedule_row, origin["iso2"])

            mark_latest(
                db,
                UsDutyCountryEffectiveRate,
                _effective_rate_filter(
                    schedule_row["hts_number"],
                    origin["iso2"],
                    assumption,
                ),
            )
            db.add(
                UsDutyCountryEffectiveRate(
                    hts_number=schedule_row["hts_number"],
                    origin_country=origin["name"],
                    origin_iso2=origin["iso2"],
                    ntr_rate_pct=schedule_row["ntr_rate_pct"],
                    fta_rate_pct=fta_rate,
                    fta_program=fta_program,
                    section_301_additional_pct=section_301,
                    ieepa_additional_pct=ieepa,
                    effective_rate_pct=effective_rate,
                    effective_rate_notes=notes,
                    yarn_forward_required=yarn_forward,
                    yarn_forward_met_assumption=assumption,
                    uflpa_risk=uflpa,
                    as_of_date=EFFECTIVE_DATE,
                    source=SOURCE_NAME,
                    is_latest=True,
                )
            )
            ctx.inserted()
            inserted += 1

    db.commit()
    return inserted


def ingest_us_duty_rates(db: Session, ctx: IngestionContext) -> dict:
    workbook_path = resolve_hts_workbook_path()
    logger.info("Loading HTS workbook from %s", workbook_path)

    schedule_rows = load_chapter_61_schedule_rows(workbook_path)
    if not schedule_rows:
        raise RuntimeError("No Chapter 61 schedule rows found in HTS workbook")

    schedule_count = ingest_schedule(db, ctx, schedule_rows)
    numeric_count = sum(1 for row in schedule_rows if row["ntr_rate_pct"] is not None)
    effective_count = ingest_effective_rates(db, ctx, schedule_rows)

    ctx.set_as_of_date(EFFECTIVE_DATE)
    logger.info(
        "Ingested %s schedule rows (%s numeric NTR), %s effective-rate rows",
        schedule_count,
        numeric_count,
        effective_count,
    )
    return {
        "schedule_rows": schedule_count,
        "numeric_ntr_rows": numeric_count,
        "effective_rate_rows": effective_count,
    }


def main() -> None:
    db = SessionLocal()
    try:
        with IngestionContext(
            source_name=SOURCE_NAME,
            script_version=SCRIPT_VERSION,
            data_source_url=DATA_SOURCE_URL,
            db=db,
        ) as ctx:
            stats = ingest_us_duty_rates(db, ctx)
            logger.info("Done: %s", stats)
    finally:
        db.close()


if __name__ == "__main__":
    main()
